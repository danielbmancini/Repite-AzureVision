import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl import load_workbook
import sys
import re

# uso: python excel4.py (linha) (nome arquivo) (nome da folha MODELO) (nome func.)


file_name = sys.argv[2] #nome do arquivo
workbook = openpyxl.load_workbook(file_name)
sheet_name = sys.argv[3] #Leva para a folha MODELO. artefato de código adaptado do outro repositório
sheet = workbook[sys.argv[3]]
sheet.title = sys.argv[4] #muda o nome da folha MODELO para o nome do funcionário
sheet["D4"] = sys.argv[4] #muda o campo de 'nome do funcionário' da planilha


# Colando a tabela na linha em que os dados começam no ponto
start_cell = f"D{7 + int(sys.argv[1]) - 1}"

#Arquivo de 4 pontos (decidir se incluir horas extras bugantes)
values_file_path = 'fin4pontos.txt'


def set_cell_value_as_number(val):
    """Sets the value of a specific cell in a specific sheet of an Excel file as a number."""
    try:
        return float(val)  # Tenta retornar número p/ conversão em tempo
    except ValueError:
        return val

#Inserir dados na planilha
def paste_table_to_excel(start_cell, values):

    # Interpreta os valores como tabela
    rows = [line.split() for line in values.strip().split('\n')]

    # Determina a célula para começar a colar (a partir da linha dada)
    start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)

   
    # Escreve os valores na planilha

    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            cell = sheet.cell(row=start_row + i, column=start_col +
                              j, value=set_cell_value_as_number(val))
    

    # Salvar
    workbook.save(file_name)

#Abrir os valores de fin4pontos
def load_values_from_file(file_path):
    with open(file_path, 'r') as file:
        return file.read()



values = load_values_from_file(values_file_path)

paste_table_to_excel(start_cell, values)
